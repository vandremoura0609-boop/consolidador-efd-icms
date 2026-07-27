from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import os
from secrets import compare_digest
from typing import Iterable

from flask import Flask, Response, render_template_string, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.datastructures import FileStorage


# -----------------------------------------------------------------------------
# Configurações gerais
# -----------------------------------------------------------------------------

APP_TITLE = "Consolidador EFD ICMS IPI"
LIMITE_UPLOAD_MB = 300
UF_PROCESSADA = "MG"
APP_USERNAME = os.environ.get("APP_USERNAME", "efd")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")

CODIGO_COMPLEMENTO = "MG109999"
CODIGO_RESSARCIMENTO = "MG120001"
CODIGO_CREDITO_PAGAMENTO_MAIOR = "MG149999"

CODIGOS_AJUSTE = {
    CODIGO_COMPLEMENTO,
    CODIGO_RESSARCIMENTO,
    CODIGO_CREDITO_PAGAMENTO_MAIOR,
}

ZERO = Decimal("0.00")

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = LIMITE_UPLOAD_MB * 1024 * 1024


# -----------------------------------------------------------------------------
# Interface web
# -----------------------------------------------------------------------------

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ app_title }}</title>
    <style>
        :root {
            --card: #ffffff;
            --primary: #2563eb;
            --primary-dark: #1d4ed8;
            --text: #111827;
            --muted: #6b7280;
            --border: #e5e7eb;
            --error-bg: #fee2e2;
            --error-border: #fecaca;
            --error-text: #991b1b;
        }

        * { box-sizing: border-box; }

        body {
            margin: 0;
            min-height: 100vh;
            font-family: Arial, Helvetica, sans-serif;
            background: linear-gradient(135deg, #f8fafc, #eef2ff);
            color: var(--text);
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 32px;
        }

        .container {
            width: 100%;
            max-width: 760px;
            background: var(--card);
            border: 1px solid var(--border);
            border-radius: 24px;
            box-shadow: 0 20px 50px rgba(15, 23, 42, 0.08);
            padding: 40px;
        }

        .badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 999px;
            background: #dbeafe;
            color: #1e40af;
            font-size: 13px;
            font-weight: 700;
            margin-bottom: 18px;
        }

        h1 {
            margin: 0 0 12px;
            font-size: 32px;
            letter-spacing: -0.04em;
        }

        p {
            margin: 0 0 28px;
            color: var(--muted);
            line-height: 1.6;
        }

        .upload-area {
            border: 2px dashed #c7d2fe;
            background: #f8fafc;
            border-radius: 20px;
            padding: 32px;
            text-align: center;
            margin-bottom: 24px;
        }

        input[type="file"] { display: none; }

        .file-label {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 14px 22px;
            border-radius: 14px;
            background: var(--primary);
            color: #ffffff;
            cursor: pointer;
            font-weight: 700;
            transition: 0.2s;
        }

        .file-label:hover { background: var(--primary-dark); }

        .selected {
            margin-top: 16px;
            font-size: 14px;
            color: var(--muted);
        }

        button {
            width: 100%;
            border: none;
            border-radius: 16px;
            background: #111827;
            color: white;
            padding: 16px 20px;
            font-size: 16px;
            font-weight: 700;
            cursor: pointer;
            transition: 0.2s;
        }

        button:hover { background: #030712; }

        button:disabled {
            opacity: 0.45;
            cursor: not-allowed;
        }

        .footer {
            margin-top: 20px;
            font-size: 13px;
            color: var(--muted);
            text-align: center;
        }

        .error {
            color: var(--error-text);
            background: var(--error-bg);
            border: 1px solid var(--error-border);
            padding: 12px 16px;
            border-radius: 12px;
            margin-bottom: 20px;
            line-height: 1.45;
        }

        @media (max-width: 640px) {
            body { padding: 16px; }
            .container { padding: 26px; }
            h1 { font-size: 26px; }
        }
    </style>
</head>
<body>
    <main class="container">
        <span class="badge">EFD ICMS IPI</span>

        <h1>Consolidador E200, E210 e E220</h1>

        <p>
            Selecione uma pasta contendo arquivos TXT da EFD ICMS IPI.
            A solução consolidará filial, competência, ajustes,
            imposto a recolher e saldo credor em uma planilha Excel.
        </p>

        {% if error %}
            <div class="error">{{ error }}</div>
        {% endif %}

        <form action="{{ url_for('processar') }}" method="post" enctype="multipart/form-data">
            <div class="upload-area">
                <label for="files" class="file-label">
                    Selecionar pasta com arquivos TXT
                </label>

                <input
                    id="files"
                    name="files"
                    type="file"
                    multiple
                    webkitdirectory
                    directory
                    accept=".txt,text/plain"
                >

                <div class="selected" id="selected">
                    Nenhum arquivo selecionado.
                </div>
            </div>

            <button type="submit" id="submit" disabled>
                Gerar planilha consolidada
            </button>
        </form>

        <div class="footer">
            Registros processados: 0000, E200, E210 e E220 — UF {{ uf_processada }}.
        </div>
    </main>

    <script>
        const input = document.getElementById("files");
        const selected = document.getElementById("selected");
        const submit = document.getElementById("submit");

        input.addEventListener("change", () => {
            const txtFiles = Array.from(input.files)
                .filter(file => file.name.toLowerCase().endsWith(".txt"));

            if (txtFiles.length === 0) {
                selected.textContent = "Nenhum arquivo TXT selecionado.";
                submit.disabled = true;
                return;
            }

            selected.textContent = `${txtFiles.length} arquivo(s) TXT selecionado(s).`;
            submit.disabled = false;
        });
    </script>
</body>
</html>
"""


# -----------------------------------------------------------------------------
# Modelos de dados
# -----------------------------------------------------------------------------

@dataclass(slots=True)
class EfdConsolidado:
    filial: str
    competencia: str
    complemento: Decimal = ZERO
    ressarcimento: Decimal = ZERO
    credito_pagamento_maior: Decimal = ZERO
    imposto_recolher: Decimal = ZERO
    saldo_credor: Decimal = ZERO
    arquivo_origem: str = ""


@dataclass(slots=True)
class E200Contexto:
    filial: str
    competencia: str
    arquivo_origem: str
    ajustes: dict[str, Decimal] = field(default_factory=dict)
    imposto_recolher: Decimal = ZERO
    saldo_credor: Decimal = ZERO


# -----------------------------------------------------------------------------
# Leitura e interpretação da EFD
# -----------------------------------------------------------------------------

def split_registro(linha: str) -> list[str]:
    """Divide uma linha da EFD preservando os índices do leiaute SPED."""
    return linha.strip().split("|")


def campo(partes: list[str], indice: int, padrao: str = "") -> str:
    """Obtém um campo pelo índice sem gerar IndexError."""
    if 0 <= indice < len(partes):
        return partes[indice].strip()
    return padrao


def parse_valor(valor: str) -> Decimal:
    """Converte valores da EFD em Decimal, aceitando vírgula ou ponto decimal."""
    texto = valor.strip().replace(" ", "")
    if not texto:
        return ZERO

    # O SPED normalmente usa vírgula como separador decimal. Este tratamento
    # também aceita arquivos que eventualmente venham com ponto decimal.
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return Decimal(texto)
    except InvalidOperation:
        return ZERO


def parse_competencia(data_inicial: str, data_final: str = "") -> str:
    """Converte DDMMAAAA para MM/AAAA."""
    data = data_inicial.strip() or data_final.strip()
    if not data:
        return ""

    try:
        return datetime.strptime(data, "%d%m%Y").strftime("%m/%Y")
    except ValueError:
        return ""


def extrair_filial_de_cnpj(cnpj: str) -> str:
    """
    Extrai o estabelecimento do CNPJ e converte para o padrão interno da filial.

    Exemplo:
        estabelecimento 0025 -> filial 5025

    A regra mantém o último trio do estabelecimento e utiliza o prefixo 5,
    reproduzindo o comportamento esperado pela solução original.
    """
    somente_digitos = "".join(caractere for caractere in cnpj if caractere.isdigit())
    if len(somente_digitos) != 14:
        return ""

    estabelecimento = somente_digitos[8:12]
    return f"5{estabelecimento[-3:]}"


def decodificar_arquivo(conteudo: bytes) -> str:
    """Tenta as codificações mais comuns em arquivos EFD."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return conteudo.decode(encoding)
        except UnicodeDecodeError:
            continue

    return conteudo.decode("latin-1", errors="replace")


def finalizar_contexto(contexto: E200Contexto | None) -> EfdConsolidado | None:
    if contexto is None:
        return None

    return EfdConsolidado(
        filial=contexto.filial,
        competencia=contexto.competencia,
        complemento=contexto.ajustes.get(CODIGO_COMPLEMENTO, ZERO),
        ressarcimento=contexto.ajustes.get(CODIGO_RESSARCIMENTO, ZERO),
        credito_pagamento_maior=contexto.ajustes.get(
            CODIGO_CREDITO_PAGAMENTO_MAIOR,
            ZERO,
        ),
        imposto_recolher=contexto.imposto_recolher,
        saldo_credor=contexto.saldo_credor,
        arquivo_origem=contexto.arquivo_origem,
    )


def processar_texto_efd(texto: str, nome_arquivo: str) -> list[EfdConsolidado]:
    """Processa um arquivo EFD e devolve os registros consolidados de MG."""
    registros: list[EfdConsolidado] = []
    filial = ""
    contexto_atual: E200Contexto | None = None

    for numero_linha, linha_original in enumerate(texto.splitlines(), start=1):
        linha = linha_original.strip()
        if not linha:
            continue

        partes = split_registro(linha)
        registro = campo(partes, 1)

        if registro == "0000":
            filial = extrair_filial_de_cnpj(campo(partes, 7))
            continue

        if registro == "E200":
            registro_finalizado = finalizar_contexto(contexto_atual)
            if registro_finalizado is not None:
                registros.append(registro_finalizado)

            uf = campo(partes, 2).upper()
            if uf == UF_PROCESSADA:
                contexto_atual = E200Contexto(
                    filial=filial,
                    competencia=parse_competencia(
                        campo(partes, 3),
                        campo(partes, 4),
                    ),
                    arquivo_origem=nome_arquivo,
                )
            else:
                contexto_atual = None
            continue

        if contexto_atual is None:
            continue

        if registro == "E210":
            contexto_atual.imposto_recolher = parse_valor(campo(partes, 13))
            contexto_atual.saldo_credor = parse_valor(campo(partes, 14))
            continue

        if registro == "E220":
            codigo_ajuste = campo(partes, 2).upper()
            if codigo_ajuste not in CODIGOS_AJUSTE:
                continue

            valor_ajuste = parse_valor(campo(partes, 4))
            contexto_atual.ajustes[codigo_ajuste] = (
                contexto_atual.ajustes.get(codigo_ajuste, ZERO) + valor_ajuste
            )
            continue

        # Mantém a variável disponível para facilitar futuros logs por linha.
        _ = numero_linha

    registro_finalizado = finalizar_contexto(contexto_atual)
    if registro_finalizado is not None:
        registros.append(registro_finalizado)

    return registros


# -----------------------------------------------------------------------------
# Geração do Excel
# -----------------------------------------------------------------------------

def criar_planilha(registros: Iterable[EfdConsolidado]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    cabecalhos = [
        "Filial",
        "Competência",
        "Complemento",
        "Ressarcimento",
        "Crédito Pagamento a Maior",
        "Imposto a Recolher",
        "Saldo Credor",
    ]
    ws.append(cabecalhos)

    registros_ordenados = sorted(
        registros,
        key=lambda item: (item.filial, item.competencia, item.arquivo_origem),
    )

    for registro in registros_ordenados:
        ws.append(
            [
                registro.filial,
                registro.competencia,
                registro.complemento,
                registro.ressarcimento,
                registro.credito_pagamento_maior,
                registro.imposto_recolher,
                registro.saldo_credor,
            ]
        )

    aplicar_formatacao_excel(ws)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def aplicar_formatacao_excel(ws: Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D9E2F3")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    moeda_format = '#,##0.00'

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        for coluna_monetaria in range(3, 8):
            row[coluna_monetaria - 1].number_format = moeda_format

    larguras = {
        "A": 14,
        "B": 16,
        "C": 18,
        "D": 18,
        "E": 28,
        "F": 22,
        "G": 18,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for indice_linha in range(1, ws.max_row + 1):
        ws.row_dimensions[indice_linha].height = 22


# -----------------------------------------------------------------------------
# Rotas Flask
# -----------------------------------------------------------------------------

@app.before_request
def proteger_aplicacao() -> Response | None:
    """Proteção opcional por usuário e senha definidos no ambiente."""
    if not APP_PASSWORD or request.endpoint == "health":
        return None

    autenticacao = request.authorization
    credenciais_validas = bool(
        autenticacao
        and compare_digest(autenticacao.username or "", APP_USERNAME)
        and compare_digest(autenticacao.password or "", APP_PASSWORD)
    )

    if credenciais_validas:
        return None

    return Response(
        "Acesso não autorizado.",
        status=401,
        headers={"WWW-Authenticate": 'Basic realm="Consolidador EFD ICMS"'},
    )


@app.get("/health")
def health() -> tuple[str, int]:
    return "ok", 200


def renderizar_pagina(error: str | None = None) -> str:
    return render_template_string(
        HTML_TEMPLATE,
        app_title=APP_TITLE,
        uf_processada=UF_PROCESSADA,
        error=error,
    )


def arquivo_txt_valido(arquivo: FileStorage) -> bool:
    return bool(
        arquivo
        and arquivo.filename
        and arquivo.filename.lower().endswith(".txt")
    )


@app.get("/")
def index() -> str:
    return renderizar_pagina()


@app.post("/processar")
def processar() -> Response | str:
    arquivos_txt = [
        arquivo
        for arquivo in request.files.getlist("files")
        if arquivo_txt_valido(arquivo)
    ]

    if not arquivos_txt:
        return renderizar_pagina("Selecione ao menos um arquivo TXT válido.")

    registros_consolidados: list[EfdConsolidado] = []
    arquivos_com_erro: list[str] = []

    for arquivo in arquivos_txt:
        try:
            texto = decodificar_arquivo(arquivo.read())
            registros_consolidados.extend(
                processar_texto_efd(
                    texto=texto,
                    nome_arquivo=arquivo.filename,
                )
            )
        except Exception:
            # Um arquivo defeituoso não interrompe o processamento dos demais.
            arquivos_com_erro.append(arquivo.filename)

    if not registros_consolidados:
        mensagem = (
            f"Nenhum registro E200 da UF {UF_PROCESSADA} foi localizado nos "
            "arquivos selecionados. Confira se os TXT pertencem à EFD ICMS IPI."
        )
        if arquivos_com_erro:
            mensagem += " Arquivos que não puderam ser processados: " + ", ".join(
                arquivos_com_erro
            )
        return renderizar_pagina(mensagem)

    planilha = criar_planilha(registros_consolidados)

    return send_file(
        planilha,
        as_attachment=True,
        download_name="consolidado_efd_icms_ipi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.errorhandler(RequestEntityTooLarge)
def tratar_arquivo_grande(_: RequestEntityTooLarge) -> tuple[str, int]:
    return (
        renderizar_pagina(
            f"O envio ultrapassou o limite de {LIMITE_UPLOAD_MB} MB. "
            "Divida os arquivos em mais de um processamento."
        ),
        413,
    )


if __name__ == "__main__":
    porta = int(os.environ.get("PORT", "5000"))
    print("=" * 65)
    print(f" {APP_TITLE}")
    print(f" Acesse no navegador: http://127.0.0.1:{porta}")
    print(" Para encerrar, pressione CTRL+C nesta janela.")
    print("=" * 65)
    app.run(host="0.0.0.0", port=porta, debug=False)
